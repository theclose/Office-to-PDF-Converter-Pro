"""
Excel Tools - Core operations for Excel file manipulation.

Functions:
- split_excel: Export each sheet as separate Excel file
- merge_excel: Combine multiple Excel files into one

Dependencies:
- openpyxl>=3.1.0

Usage:
    from office_converter.core.excel_tools import split_excel, merge_excel
    
    # Split Excel file
    output_files = split_excel("report.xlsx")
    
    # Merge Excel files
    output_path = merge_excel(["file1.xlsx", "file2.xlsx"], "merged.xlsx")
"""

import logging
from pathlib import Path
from typing import List, Optional, Callable, Tuple
from copy import copy

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


# ============================================================================
# SPLIT EXCEL
# ============================================================================

def split_excel(
    input_path: str,
    output_dir: Optional[str] = None,
    sheets: Optional[List[str]] = None,
    on_progress: Optional[Callable[[int, int, str], None]] = None
) -> Tuple[List[str], List[str]]:
    """
    Split Excel file - export each sheet as separate Excel file.
    
    Args:
        input_path: Path to source Excel file (.xlsx)
        output_dir: Output directory (default: same as input file)
        sheets: List of specific sheet names to export (None = all sheets)
        on_progress: Callback(current, total, message) for progress updates
    
    Returns:
        Tuple of (success_files, error_messages)
        
    Output naming: {original_name}_{sheet_name}.xlsx
    
    Example:
        >>> files, errors = split_excel("report.xlsx")
        >>> print(f"Created {len(files)} files")
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel operations. Install: pip install openpyxl")
    
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")
    
    if input_path.suffix.lower() not in ['.xlsx', '.xlsm']:
        raise ValueError(f"Unsupported format: {input_path.suffix}. Only .xlsx/.xlsm supported.")
    
    # Determine output directory
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = input_path.parent
    
    base_name = input_path.stem
    success_files = []
    errors = []
    
    try:
        # Load workbook (data_only=False to preserve formulas)
        wb = load_workbook(str(input_path), data_only=False)
        sheet_names = sheets if sheets else wb.sheetnames
        total = len(sheet_names)
        
        logger.info(f"Splitting '{input_path.name}' into {total} sheet(s)")
        
        for idx, sheet_name in enumerate(sheet_names):
            try:
                if sheet_name not in wb.sheetnames:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue
                
                # Report progress
                if on_progress:
                    on_progress(idx + 1, total, f"Exporting: {sheet_name}")
                
                # Create new workbook with single sheet
                new_wb = Workbook()
                new_ws = new_wb.active
                new_ws.title = sheet_name
                
                # Copy data from source sheet
                src_ws = wb[sheet_name]
                _copy_sheet_data(src_ws, new_ws)
                
                # Generate output path (sanitize sheet name for filename)
                safe_sheet_name = _sanitize_filename(sheet_name)
                output_path = output_dir / f"{safe_sheet_name}.xlsx"
                
                # Handle duplicate filenames
                counter = 1
                while output_path.exists():
                    output_path = output_dir / f"{safe_sheet_name}_{counter}.xlsx"
                    counter += 1
                
                # Save
                new_wb.save(str(output_path))
                new_wb.close()
                
                success_files.append(str(output_path))
                logger.info(f"✅ Created: {output_path.name}")
                
            except Exception as e:
                error_msg = f"Error exporting sheet '{sheet_name}': {e}"
                errors.append(error_msg)
                logger.error(error_msg)
        
        wb.close()
        
    except Exception as e:
        errors.append(f"Failed to open file: {e}")
        logger.error(f"Split failed: {e}")
    
    return success_files, errors


# ============================================================================
# MERGE EXCEL
# ============================================================================

def merge_excel(
    input_files: List[str],
    output_path: str,
    mode: str = "sheets",
    skip_header_after_first: bool = True,
    on_progress: Optional[Callable[[int, int, str], None]] = None
) -> Tuple[str, List[str]]:
    """
    Merge multiple Excel files into one.
    
    Args:
        input_files: List of Excel file paths (.xlsx)
        output_path: Output file path
        mode: Merge mode
            - "sheets": Each source sheet becomes a sheet in output
            - "rows": Append all data into a single sheet
        skip_header_after_first: (rows mode) Skip header row for files after first
        on_progress: Callback(current, total, message) for progress updates
    
    Returns:
        Tuple of (output_path, error_messages)
        
    Example:
        >>> path, errors = merge_excel(["q1.xlsx", "q2.xlsx"], "annual.xlsx")
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel operations. Install: pip install openpyxl")
    
    if not input_files:
        raise ValueError("No input files provided")
    
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    errors = []
    
    try:
        if mode == "sheets":
            return _merge_as_sheets(input_files, output_path, on_progress, errors)
        elif mode == "rows":
            return _merge_as_rows(input_files, output_path, skip_header_after_first, on_progress, errors)
        else:
            raise ValueError(f"Unknown mode: {mode}. Use 'sheets' or 'rows'.")
    
    except Exception as e:
        errors.append(f"Merge failed: {e}")
        logger.error(f"Merge error: {e}")
        return "", errors


def _merge_as_sheets(
    input_files: List[str],
    output_path: Path,
    on_progress: Optional[Callable],
    errors: List[str]
) -> Tuple[str, List[str]]:
    """Merge files - each sheet becomes a sheet in output."""
    
    merged_wb = Workbook()
    # Remove default sheet (will add actual sheets)
    default_sheet = merged_wb.active
    
    total = len(input_files)
    sheet_count = 0
    used_names = set()
    
    for file_idx, file_path in enumerate(input_files):
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                errors.append(f"File not found: {file_path}")
                continue
            
            if on_progress:
                on_progress(file_idx + 1, total, f"Processing: {file_path.name}")
            
            src_wb = load_workbook(str(file_path), data_only=False)
            file_base = file_path.stem
            
            for sheet_name in src_wb.sheetnames:
                # Generate unique sheet name
                new_name = f"{file_base}_{sheet_name}"
                new_name = new_name[:31]  # Excel limit: 31 chars
                new_name = _make_unique_name(new_name, used_names)
                used_names.add(new_name)
                
                # Create new sheet
                new_ws = merged_wb.create_sheet(title=new_name)
                
                # Copy data
                src_ws = src_wb[sheet_name]
                _copy_sheet_data(src_ws, new_ws)
                
                sheet_count += 1
            
            src_wb.close()
            logger.info(f"✅ Added sheets from: {file_path.name}")
            
        except Exception as e:
            error_msg = f"Error processing {file_path}: {e}"
            errors.append(error_msg)
            logger.error(error_msg)
    
    # Remove default empty sheet if we added other sheets
    if sheet_count > 0 and default_sheet.title in merged_wb.sheetnames:
        merged_wb.remove(default_sheet)
    
    # Save
    merged_wb.save(str(output_path))
    merged_wb.close()
    
    logger.info(f"✅ Merged {sheet_count} sheets into: {output_path.name}")
    
    return str(output_path), errors


def _merge_as_rows(
    input_files: List[str],
    output_path: Path,
    skip_header: bool,
    on_progress: Optional[Callable],
    errors: List[str]
) -> Tuple[str, List[str]]:
    """Merge files - append all rows into single sheet."""
    
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "Merged"
    
    total = len(input_files)
    current_row = 1
    is_first_file = True
    
    for file_idx, file_path in enumerate(input_files):
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                errors.append(f"File not found: {file_path}")
                continue
            
            if on_progress:
                on_progress(file_idx + 1, total, f"Processing: {file_path.name}")
            
            src_wb = load_workbook(str(file_path), data_only=True, read_only=True)
            
            for sheet_name in src_wb.sheetnames:
                src_ws = src_wb[sheet_name]
                
                for row_idx, row in enumerate(src_ws.iter_rows(values_only=True)):
                    # Skip header for subsequent files if enabled
                    if skip_header and row_idx == 0 and not is_first_file:
                        continue
                    
                    # Write row
                    for col_idx, value in enumerate(row, start=1):
                        merged_ws.cell(row=current_row, column=col_idx, value=value)
                    
                    current_row += 1
                
                is_first_file = False
            
            src_wb.close()
            logger.info(f"✅ Appended rows from: {file_path.name}")
            
        except Exception as e:
            error_msg = f"Error processing {file_path}: {e}"
            errors.append(error_msg)
            logger.error(error_msg)
    
    # Save
    merged_wb.save(str(output_path))
    merged_wb.close()
    
    logger.info(f"✅ Merged {current_row - 1} rows into: {output_path.name}")
    
    return str(output_path), errors


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def _copy_sheet_data(src_ws: 'Worksheet', dst_ws: 'Worksheet'):
    """Copy data and basic formatting from source to destination worksheet."""
    
    # Copy cell values and basic formatting
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            
            # Copy basic formatting if exists
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
    
    # Copy column widths
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
    
    # Copy row heights
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height
    
    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))


def _sanitize_filename(name: str) -> str:
    """Sanitize string for use in filename."""
    # Remove/replace invalid characters
    invalid_chars = '<>:"/\\|?*'
    result = name
    for char in invalid_chars:
        result = result.replace(char, '_')
    return result.strip()


def _make_unique_name(name: str, used_names: set) -> str:
    """Make name unique by adding suffix if needed."""
    if name not in used_names:
        return name
    
    counter = 1
    while f"{name}_{counter}" in used_names:
        counter += 1
    
    return f"{name}_{counter}"


def get_sheet_names(file_path: str) -> List[str]:
    """Get list of sheet names from Excel file.
    
    Useful for UI to show available sheets for split operation.
    """
    if not HAS_OPENPYXL:
        return []
    
    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        logger.error(f"Error reading sheets: {e}")
        return []


def get_sheet_info(file_path: str) -> List[dict]:
    """Get detailed info about sheets in Excel file.
    
    Returns list of dicts with:
    - name: Sheet name
    - rows: Number of rows
    - cols: Number of columns
    """
    if not HAS_OPENPYXL:
        return []
    
    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        info = []
        
        for name in wb.sheetnames:
            ws = wb[name]
            info.append({
                'name': name,
                'rows': ws.max_row or 0,
                'cols': ws.max_column or 0,
            })
        
        wb.close()
        return info
    except Exception as e:
        logger.error(f"Error reading sheet info: {e}")
        return []


# ============================================================================
# PHASE 1: EXCEL ↔ CSV CONVERSION
# ============================================================================

import csv

def excel_to_csv(
    input_path: str,
    output_dir: Optional[str] = None,
    sheets: Optional[List[str]] = None,
    encoding: str = 'utf-8-sig',
    delimiter: str = ',',
    on_progress: Optional[Callable[[int, int, str], None]] = None
) -> Tuple[List[str], List[str]]:
    """
    Export Excel sheets to CSV files.
    
    Args:
        input_path: Path to Excel file (.xlsx)
        output_dir: Output directory (default: same as input)
        sheets: Specific sheets to export (None = all)
        encoding: Output encoding (utf-8-sig for Excel compatibility, utf-8, utf-16)
        delimiter: CSV delimiter (,;|\t)
        on_progress: Progress callback
    
    Returns:
        Tuple of (success_files, errors)
        
    Output: {filename}_{sheetname}.csv
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required")
    
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")
    
    output_dir = Path(output_dir) if output_dir else input_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)
    
    success_files = []
    errors = []
    base_name = input_path.stem
    
    try:
        wb = load_workbook(str(input_path), read_only=True, data_only=True)
        sheet_names = sheets if sheets else wb.sheetnames
        total = len(sheet_names)
        
        for idx, sheet_name in enumerate(sheet_names):
            try:
                if sheet_name not in wb.sheetnames:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue
                
                if on_progress:
                    on_progress(idx + 1, total, f"Exporting: {sheet_name}")
                
                ws = wb[sheet_name]
                safe_name = _sanitize_filename(sheet_name)
                output_path = output_dir / f"{base_name}_{safe_name}.csv"
                
                with open(output_path, 'w', newline='', encoding=encoding) as f:
                    writer = csv.writer(f, delimiter=delimiter)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
                
                success_files.append(str(output_path))
                logger.info(f"✅ Created CSV: {output_path.name}")
                
            except Exception as e:
                errors.append(f"Error exporting '{sheet_name}': {e}")
                logger.error(f"CSV export error: {e}")
        
        wb.close()
        
    except Exception as e:
        errors.append(f"Failed to open file: {e}")
        logger.error(f"Excel to CSV failed: {e}")
    
    return success_files, errors


def csv_to_excel(
    input_files: List[str],
    output_path: str,
    encoding: str = 'utf-8-sig',
    delimiter: str = ',',
    on_progress: Optional[Callable[[int, int, str], None]] = None
) -> Tuple[str, List[str]]:
    """
    Import CSV files into Excel workbook.
    
    Args:
        input_files: List of CSV file paths
        output_path: Output Excel file path
        encoding: CSV encoding
        delimiter: CSV delimiter
        on_progress: Progress callback
    
    Returns:
        Tuple of (output_path, errors)
        
    Each CSV becomes a sheet named after the file.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required")
    
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    errors = []
    wb = Workbook()
    default_sheet = wb.active
    used_names = set()
    sheet_count = 0
    total = len(input_files)
    
    for idx, csv_path in enumerate(input_files):
        try:
            csv_path = Path(csv_path)
            if not csv_path.exists():
                errors.append(f"File not found: {csv_path}")
                continue
            
            if on_progress:
                on_progress(idx + 1, total, f"Importing: {csv_path.name}")
            
            # Create sheet
            sheet_name = csv_path.stem[:31]  # Excel limit
            sheet_name = _make_unique_name(sheet_name, used_names)
            used_names.add(sheet_name)
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Read CSV and write to sheet
            with open(csv_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row_idx, row in enumerate(reader, start=1):
                    for col_idx, value in enumerate(row, start=1):
                        # Try to convert to number if possible
                        try:
                            if '.' in value:
                                value = float(value)
                            else:
                                value = int(value)
                        except (ValueError, TypeError):
                            pass
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            sheet_count += 1
            logger.info(f"✅ Imported CSV: {csv_path.name}")
            
        except Exception as e:
            errors.append(f"Error importing '{csv_path}': {e}")
            logger.error(f"CSV import error: {e}")
    
    # Remove default sheet if we added others
    if sheet_count > 0 and default_sheet.title in wb.sheetnames:
        wb.remove(default_sheet)
    
    wb.save(str(output_path))
    wb.close()
    
    logger.info(f"✅ Created Excel with {sheet_count} sheets")
    return str(output_path), errors


# ============================================================================
# PHASE 1: SHEET PROTECTION
# ============================================================================

def protect_sheets(
    input_path: str,
    output_path: Optional[str] = None,
    password: str = "",
    sheets: Optional[List[str]] = None,
    on_progress: Optional[Callable[[int, int, str], None]] = None
) -> Tuple[str, List[str]]:
    """
    Protect sheets with password.
    
    Args:
        input_path: Excel file path
        output_path: Output path (None = overwrite input)
        password: Protection password (empty = protect without password)
        sheets: Specific sheets (None = all)
        on_progress: Progress callback
    
    Returns:
        Tuple of (output_path, errors)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required")
    
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")
    
    output_path = Path(output_path) if output_path else input_path
    errors = []
    
    try:
        wb = load_workbook(str(input_path))
        sheet_names = sheets if sheets else wb.sheetnames
        total = len(sheet_names)
        
        for idx, sheet_name in enumerate(sheet_names):
            try:
                if sheet_name not in wb.sheetnames:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue
                
                if on_progress:
                    on_progress(idx + 1, total, f"Protecting: {sheet_name}")
                
                ws = wb[sheet_name]
                ws.protection.sheet = True
                ws.protection.password = password
                
                logger.info(f"✅ Protected: {sheet_name}")
                
            except Exception as e:
                errors.append(f"Error protecting '{sheet_name}': {e}")
        
        wb.save(str(output_path))
        wb.close()
        
    except Exception as e:
        errors.append(f"Failed: {e}")
        logger.error(f"Protect failed: {e}")
        return "", errors
    
    return str(output_path), errors


def unprotect_sheets(
    input_path: str,
    output_path: Optional[str] = None,
    password: str = "",
    sheets: Optional[List[str]] = None,
    on_progress: Optional[Callable[[int, int, str], None]] = None
) -> Tuple[str, List[str]]:
    """
    Remove protection from sheets.
    
    Args:
        input_path: Excel file path
        output_path: Output path (None = overwrite)
        password: Current protection password
        sheets: Specific sheets (None = all)
        on_progress: Progress callback
    
    Returns:
        Tuple of (output_path, errors)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required")
    
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")
    
    output_path = Path(output_path) if output_path else input_path
    errors = []
    
    try:
        wb = load_workbook(str(input_path))
        sheet_names = sheets if sheets else wb.sheetnames
        total = len(sheet_names)
        
        for idx, sheet_name in enumerate(sheet_names):
            try:
                if sheet_name not in wb.sheetnames:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue
                
                if on_progress:
                    on_progress(idx + 1, total, f"Unprotecting: {sheet_name}")
                
                ws = wb[sheet_name]
                ws.protection.sheet = False
                ws.protection.password = None
                
                logger.info(f"✅ Unprotected: {sheet_name}")
                
            except Exception as e:
                errors.append(f"Error unprotecting '{sheet_name}': {e}")
        
        wb.save(str(output_path))
        wb.close()
        
    except Exception as e:
        errors.append(f"Failed: {e}")
        logger.error(f"Unprotect failed: {e}")
        return "", errors
    
    return str(output_path), errors


# ============================================================================
# PHASE 1: BATCH RENAME SHEETS
# ============================================================================

def rename_sheets(
    input_path: str,
    output_path: Optional[str] = None,
    rename_map: Optional[dict] = None,
    prefix: str = "",
    suffix: str = "",
    replace_from: str = "",
    replace_to: str = "",
    on_progress: Optional[Callable[[int, int, str], None]] = None
) -> Tuple[str, List[str]]:
    """
    Rename sheets in Excel file.
    
    Args:
        input_path: Excel file path
        output_path: Output path (None = overwrite)
        rename_map: Dict of {old_name: new_name} for specific renames
        prefix: Add prefix to all sheet names
        suffix: Add suffix to all sheet names
        replace_from: Text to replace in sheet names
        replace_to: Replacement text
        on_progress: Progress callback
    
    Returns:
        Tuple of (output_path, errors)
        
    Priority: rename_map > prefix/suffix > replace
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required")
    
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")
    
    output_path = Path(output_path) if output_path else input_path
    errors = []
    renamed_count = 0
    
    try:
        wb = load_workbook(str(input_path))
        total = len(wb.sheetnames)
        used_names = set()
        
        for idx, sheet_name in enumerate(wb.sheetnames):
            try:
                if on_progress:
                    on_progress(idx + 1, total, f"Processing: {sheet_name}")
                
                ws = wb[sheet_name]
                new_name = sheet_name
                
                # Apply rename map first
                if rename_map and sheet_name in rename_map:
                    new_name = rename_map[sheet_name]
                else:
                    # Apply prefix/suffix
                    if prefix:
                        new_name = prefix + new_name
                    if suffix:
                        new_name = new_name + suffix
                    
                    # Apply replace
                    if replace_from:
                        new_name = new_name.replace(replace_from, replace_to)
                
                # Truncate to Excel limit
                new_name = new_name[:31]
                
                # Make unique
                if new_name != sheet_name:
                    new_name = _make_unique_name(new_name, used_names)
                    ws.title = new_name
                    renamed_count += 1
                    logger.info(f"✅ Renamed: {sheet_name} → {new_name}")
                
                used_names.add(new_name)
                
            except Exception as e:
                errors.append(f"Error renaming '{sheet_name}': {e}")
        
        wb.save(str(output_path))
        wb.close()
        
        logger.info(f"✅ Renamed {renamed_count} sheets")
        
    except Exception as e:
        errors.append(f"Failed: {e}")
        logger.error(f"Rename failed: {e}")
        return "", errors
    
    return str(output_path), errors

